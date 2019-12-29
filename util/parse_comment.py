# -*- coding: utf-8 -*-

import openpyxl
from openpyxl import load_workbook

xl_file = "D:\\work\\Python\\PN3\\history\\transactions.xlsx";
comment_column_number = 9
XLList = 'List1'             # Лист с транзакциями



wb = load_workbook(xl_file)

#print(wb.get_sheet_names())

sheet = wb.get_sheet_by_name(XLList)

#print(sheet['I2'].value)



for r in range(2,sheet.max_row):
    cellV = sheet.cell(row=r, column=comment_column_number).value
    splittedcell = cellV.split(" ")
    if(splittedcell[0]=='Sold' or splittedcell[0]=='Bought'):
        sheet.cell(row=r, column=comment_column_number+1).value = (splittedcell[1]).replace('.',',')
        sheet.cell(row=r, column=comment_column_number + 2).value = splittedcell[2]
        sheet.cell(row=r, column=comment_column_number + 3).value = splittedcell[4].replace('.',',')
        sheet.cell(row=r, column=comment_column_number + 4).value = splittedcell[5]

        if splittedcell[0]=='Sold':
            n = str(r)
            sheet.cell(row=r, column=comment_column_number + 5).value = "=H"+n+"/(J"+n+"*L"+n+")*100"


wb.save(xl_file)